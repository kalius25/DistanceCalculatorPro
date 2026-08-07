from pathlib import Path

import pytest
from openpyxl import Workbook

from app.workbooks import (
    CsvVirtualWorksheetDataSource,
    OpenPyXLVirtualWorksheetDataSource,
    UnsupportedVirtualWorkbookError,
    VirtualWorksheetDataSourceFactory,
    VirtualWorksheetNotFoundError,
)


def test_csv_virtual_source_reads_bounded_ranges(tmp_path: Path) -> None:
    path = tmp_path / "routes.csv"
    path.write_text(
        "Origin,Destination,Note\nA,B\nC,D,Fast\nE,F,\n",
        encoding="utf-8",
    )

    source = CsvVirtualWorksheetDataSource(path)

    assert source.file_path == path
    assert source.worksheet_name == "routes"
    assert source.headers == ("Origin", "Destination", "Note")
    assert source.row_count == 3
    assert source.column_count == 3
    assert source.read_rows(1, 2) == (("C", "D", "Fast"), ("E", "F", ""))
    assert source.read_rows(3, 1) == ()
    assert source.read_rows(0, 0) == ()


def test_csv_virtual_source_handles_empty_and_wide_rows(tmp_path: Path) -> None:
    empty = tmp_path / "empty.csv"
    empty.touch()
    source = CsvVirtualWorksheetDataSource(empty)

    assert source.headers == ()
    assert source.row_count == 0
    assert source.column_count == 0

    wide = tmp_path / "wide.csv"
    wide.write_text("A,B\n1,2,3\n", encoding="utf-8")
    source = CsvVirtualWorksheetDataSource(wide)

    assert source.headers == ("A", "B", "")
    assert source.read_rows(0, 1) == (("1", "2", "3"),)


def test_virtual_source_range_validation_and_close(tmp_path: Path) -> None:
    path = tmp_path / "routes.csv"
    path.write_text("A\n1\n", encoding="utf-8")
    source = CsvVirtualWorksheetDataSource(path)

    with pytest.raises(ValueError, match="start must be non-negative"):
        source.read_rows(-1, 1)
    with pytest.raises(ValueError, match="count must be non-negative"):
        source.read_rows(0, -1)

    source.close()
    with pytest.raises(RuntimeError, match="data source is closed"):
        source.read_rows(0, 1)


def test_excel_virtual_source_reads_requested_rows_and_closes(tmp_path: Path) -> None:
    path = tmp_path / "routes.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Routes"
    worksheet.append(["Origin", "Destination", "Note"])
    worksheet.append(["A", "B"])
    worksheet.append(["C", "D", "Fast"])
    worksheet.append(["E", None, "Last"])
    workbook.save(path)
    workbook.close()

    source = OpenPyXLVirtualWorksheetDataSource(path, "Routes")

    assert source.file_path == path
    assert source.worksheet_name == "Routes"
    assert source.headers == ("Origin", "Destination", "Note")
    assert source.row_count == 3
    assert source.column_count == 3
    assert source.read_rows(0, 2) == (("A", "B", ""), ("C", "D", "Fast"))
    assert source.read_rows(2, 10) == (("E", "", "Last"),)
    assert source.read_rows(3, 1) == ()
    assert source.read_rows(0, 0) == ()

    source.close()
    source.close()
    with pytest.raises(RuntimeError, match="data source is closed"):
        source.read_rows(0, 1)


def test_excel_virtual_source_validates_ranges(tmp_path: Path) -> None:
    path = tmp_path / "routes.xlsx"
    workbook = Workbook()
    workbook.active.append(["A"])
    workbook.active.append([1])
    workbook.save(path)
    workbook.close()
    source = OpenPyXLVirtualWorksheetDataSource(path, "Sheet")

    with pytest.raises(ValueError, match="start must be non-negative"):
        source.read_rows(-1, 1)
    with pytest.raises(ValueError, match="count must be non-negative"):
        source.read_rows(0, -1)
    source.close()


def test_excel_virtual_source_rejects_missing_sheet(tmp_path: Path) -> None:
    path = tmp_path / "routes.xlsx"
    workbook = Workbook()
    workbook.save(path)
    workbook.close()

    with pytest.raises(VirtualWorksheetNotFoundError, match="Worksheet not found"):
        OpenPyXLVirtualWorksheetDataSource(path, "Missing")


def test_virtual_source_factory_selects_supported_sources(tmp_path: Path) -> None:
    csv_path = tmp_path / "routes.csv"
    csv_path.write_text("A\n1\n", encoding="utf-8")
    excel_path = tmp_path / "routes.xlsx"
    workbook = Workbook()
    workbook.save(excel_path)
    workbook.close()
    factory = VirtualWorksheetDataSourceFactory()

    csv_source = factory.create(csv_path)
    assert isinstance(csv_source, CsvVirtualWorksheetDataSource)
    csv_source.close()

    excel_source = factory.create(excel_path, "Sheet")
    assert isinstance(excel_source, OpenPyXLVirtualWorksheetDataSource)
    excel_source.close()


def test_virtual_source_factory_rejects_invalid_requests(tmp_path: Path) -> None:
    factory = VirtualWorksheetDataSourceFactory()
    excel_path = tmp_path / "routes.xlsx"
    workbook = Workbook()
    workbook.save(excel_path)
    workbook.close()
    csv_path = tmp_path / "routes.csv"
    csv_path.write_text("A\n", encoding="utf-8")

    with pytest.raises(VirtualWorksheetNotFoundError, match="name is required"):
        factory.create(excel_path)
    with pytest.raises(VirtualWorksheetNotFoundError, match="Worksheet not found"):
        factory.create(csv_path, "Other")
    with pytest.raises(
        UnsupportedVirtualWorkbookError,
        match="Unsupported virtual workbook format",
    ):
        factory.create(tmp_path / "routes.txt")
