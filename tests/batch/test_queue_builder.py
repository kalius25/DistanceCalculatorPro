from pathlib import Path

from openpyxl import Workbook

from app.batch import QueueBuilder, RouteJobStatus
from app.enums.provider_type import ProviderType
from app.enums.travel_mode import TravelMode
from app.presentation.workspace_configuration import (
    ColumnMapping,
    ProviderConfiguration,
    WorkspaceConfiguration,
)


def configuration() -> WorkspaceConfiguration:
    return WorkspaceConfiguration(
        ColumnMapping("Origin", "Destination", "Distance"),
        ProviderConfiguration(
            ProviderType.GOOGLE_MAPS_WEB,
            TravelMode.DRIVING,
        ),
    )


def test_queue_builder_creates_jobs_for_all_source_rows(tmp_path: Path) -> None:
    path = tmp_path / "routes.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Routes"
    sheet.append(["Origin", "Destination", "Distance"])
    sheet.append(["A", "B", None])
    sheet.append(["", "B", None])
    sheet.append(["C", "C", None])
    sheet.append(["91,106", "10,106", None])
    workbook.save(path)
    workbook.close()

    queue = QueueBuilder().build(path, "Routes", configuration())

    assert len(queue) == 4
    assert queue.ready_count == 1
    assert queue.skipped_count == 1
    assert queue.invalid_count == 1
    assert queue.count(RouteJobStatus.DONE) == 1


def test_queue_builder_skips_existing_results_when_enabled(tmp_path: Path) -> None:
    path = tmp_path / "resume.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Routes"
    sheet.append(["Origin", "Destination", "Distance"])
    sheet.append(["A", "B", 8.6])
    sheet.append(["C", "D", None])
    sheet.append(["E", "F", "ERROR: timeout"])
    workbook.save(path)
    workbook.close()

    queue = QueueBuilder().build(path, "Routes", configuration())

    assert queue.done_count == 1
    assert queue.pending_count == 2
    resumed = next(job for job in queue if job.row_index == 2)
    assert resumed.result_distance_km == 8.6
    assert resumed.metadata["resumed_existing_result"] is True


def test_queue_builder_can_reprocess_existing_results(tmp_path: Path) -> None:
    path = tmp_path / "rerun.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Routes"
    sheet.append(["Origin", "Destination", "Distance"])
    sheet.append(["A", "B", 8.6])
    workbook.save(path)
    workbook.close()

    base = configuration()
    config = WorkspaceConfiguration(
        base.column_mapping,
        base.provider_configuration,
        skip_existing_results=False,
    )
    queue = QueueBuilder().build(path, "Routes", config)

    assert queue.pending_count == 1
    assert queue.done_count == 0
