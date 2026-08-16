from __future__ import annotations

import csv
from io import StringIO
from pathlib import Path
from unittest.mock import MagicMock, call, patch

import pytest
from openpyxl import Workbook, load_workbook

from app.batch import OutputWriteError
from app.batch.autosave_policy import AutoSavePolicy
from app.batch.models import RouteJob, RouteJobStatus
from app.batch.output_path_policy import OutputPathPolicy
from app.batch.result_writer import (
    CsvResultWriter,
    ExcelResultWriter,
    ResultWriterFactory,
)


def done_job(row: int = 2, distance: float = 8.6) -> RouteJob:
    return RouteJob(
        row,
        "A",
        "B",
        "Distance",
        status=RouteJobStatus.DONE,
        result_distance_km=distance,
    )


def test_excel_writer_writes_and_flushes_result(tmp_path: Path) -> None:
    source = tmp_path / "routes.xlsx"
    output = tmp_path / "routes.result.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Routes"
    sheet.append(["Origin", "Destination", "Distance"])
    sheet.append(["A", "B", None])
    workbook.save(source)
    workbook.close()

    policy = AutoSavePolicy(1, 60.0)
    writer = ExcelResultWriter(source, "Routes", output, policy)

    assert writer.write(done_job())
    assert output.exists()
    assert not writer.dirty
    assert not writer.flush()
    writer.close()
    assert writer.closed
    writer.close()

    saved = load_workbook(output, data_only=True)
    assert saved["Routes"]["C2"].value == 8.6
    saved.close()



def test_excel_writer_loads_source_from_memory_snapshot(tmp_path: Path) -> None:
    source = tmp_path / "routes.result.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Routes"
    sheet.append(["Origin", "Destination", "Distance"])
    sheet.append(["A", "B", None])
    workbook.save(source)
    workbook.close()

    real_load_workbook = load_workbook
    with patch(
        "app.batch.result_writer.load_workbook",
        wraps=real_load_workbook,
    ) as mocked_load:
        writer = ExcelResultWriter(source, "Routes", source)

    loaded_from = mocked_load.call_args.args[0]
    assert not isinstance(loaded_from, (str, Path))

    assert writer.write(done_job())
    writer.close()

    saved = load_workbook(source, data_only=True)
    assert saved["Routes"]["C2"].value == 8.6
    saved.close()

def test_excel_writer_preserves_errors_and_validates_inputs(
    tmp_path: Path,
) -> None:
    source = tmp_path / "routes.xlsx"
    output = tmp_path / "routes.result.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Routes"
    sheet.append(["Origin", "Destination", "Distance"])
    sheet.append(["A", "B", None])
    workbook.save(source)
    workbook.close()

    writer = ExcelResultWriter(source, "Routes", output)
    failed = RouteJob(
        2,
        "A",
        "B",
        "Distance",
        status=RouteJobStatus.FAILED,
        validation_error="timeout",
    )
    assert writer.write(failed)
    skipped = RouteJob(
        2,
        "",
        "B",
        "Distance",
        status=RouteJobStatus.SKIPPED,
    )
    assert not writer.write(skipped)
    writer.close()

    with pytest.raises(ValueError, match="Worksheet not found"):
        ExcelResultWriter(source, "Missing", output)

    writer = ExcelResultWriter(source, "Routes", output)
    with pytest.raises(ValueError, match="Result column not found"):
        writer.write(
            RouteJob(
                2,
                "A",
                "B",
                "Missing",
                status=RouteJobStatus.DONE,
                result_distance_km=1.0,
            )
        )
    writer.close()


def test_csv_writer_rewrites_output_and_validates_rows(tmp_path: Path) -> None:
    source = tmp_path / "routes.csv"
    output = tmp_path / "routes.result.csv"
    source.write_text(
        "Origin,Destination,Distance\nA,B,\n",
        encoding="utf-8",
    )
    writer = CsvResultWriter(
        source,
        output,
        AutoSavePolicy(1, 60.0),
    )
    assert writer.write(done_job())
    writer.close()

    with output.open("r", encoding="utf-8-sig", newline="") as stream:
        rows = list(csv.reader(stream))
    assert rows[1][2] == "8.6"

    writer = CsvResultWriter(source, output)
    with pytest.raises(ValueError, match="CSV row not found"):
        writer.write(done_job(row=9))
    with pytest.raises(ValueError, match="Result column not found"):
        writer.write(
            RouteJob(
                2,
                "A",
                "B",
                "Missing",
                status=RouteJobStatus.DONE,
                result_distance_km=1.0,
            )
        )
    writer.close()

    empty = tmp_path / "empty.csv"
    empty.touch()
    with pytest.raises(ValueError, match="header row"):
        CsvResultWriter(empty, output)


def test_result_writer_factory_selects_supported_writer(
    tmp_path: Path,
) -> None:
    xlsx = tmp_path / "routes.xlsx"
    xlsm = tmp_path / "routes.xlsm"
    csv_path = tmp_path / "routes.csv"

    xlsx.touch()
    xlsm.touch()
    csv_path.touch()

    output_policy = OutputPathPolicy()
    factory = ResultWriterFactory(output_policy)

    xlsx_writer = MagicMock(spec=ExcelResultWriter)
    xlsm_writer = MagicMock(spec=ExcelResultWriter)
    csv_writer = MagicMock(spec=CsvResultWriter)

    with (
        patch(
            "app.batch.result_writer.ExcelResultWriter",
            side_effect=[
                xlsx_writer,
                xlsm_writer,
            ],
        ) as excel_writer_type,
        patch(
            "app.batch.result_writer.CsvResultWriter",
            return_value=csv_writer,
        ) as csv_writer_type,
    ):
        assert factory.create(xlsx, "Routes") is xlsx_writer
        assert factory.create(xlsm, "Routes") is xlsm_writer
        assert factory.create(csv_path, "Routes") is csv_writer

    assert excel_writer_type.call_args_list == [
        call(
            xlsx,
            "Routes",
            output_policy.build(xlsx),
            None,
        ),
        call(
            xlsm,
            "Routes",
            output_policy.build(xlsm),
            None,
        ),
    ]

    csv_writer_type.assert_called_once_with(
        csv_path,
        output_policy.build(csv_path),
        None,
    )

    unsupported = tmp_path / "routes.txt"
    unsupported.touch()

    with pytest.raises(
        ValueError,
        match="Unsupported workbook type",
    ):
        factory.create(
            unsupported,
            "Routes",
        )


def test_writer_context_manager_and_unknown_error(tmp_path: Path) -> None:
    source = tmp_path / "routes.csv"
    output = tmp_path / "routes.result.csv"
    source.write_text(
        "Origin,Destination,Distance\nA,B,\n",
        encoding="utf-8",
    )
    with CsvResultWriter(source, output) as writer:
        invalid = RouteJob(
            2,
            "A",
            "B",
            "Distance",
            status=RouteJobStatus.INVALID,
        )
        assert writer.write(invalid)
    assert writer.closed


def test_csv_writer_extends_short_rows(tmp_path: Path) -> None:
    source = tmp_path / "routes.csv"
    output = tmp_path / "routes.result.csv"
    source.write_text(
        "Origin,Destination,Distance\nA,B\n",
        encoding="utf-8",
    )
    writer = CsvResultWriter(source, output, AutoSavePolicy(1, 60.0))

    assert writer.write(done_job())
    writer.close()

    with output.open("r", encoding="utf-8-sig", newline="") as stream:
        rows = list(csv.reader(stream))
    assert rows[1] == ["A", "B", "8.6"]


def test_factory_can_resume_from_existing_output(tmp_path: Path) -> None:
    source = tmp_path / "routes.csv"
    output = tmp_path / "routes.result.csv"
    source.write_text("Origin,Destination,Distance\nA,B,\n", encoding="utf-8")
    output.write_text("Origin,Destination,Distance\nA,B,8.6\n", encoding="utf-8")

    writer = ResultWriterFactory().create(
        source,
        "Routes",
        resume_from_output=True,
    )

    assert isinstance(writer, CsvResultWriter)
    assert writer._rows[1][2] == "8.6"
    writer.close()


def test_writer_records_autosave_metrics(tmp_path: Path) -> None:
    source = tmp_path / "routes.csv"
    output = tmp_path / "routes.result.csv"
    source.write_text(
        "Origin,Destination,Distance\nA,B,\n",
        encoding="utf-8",
    )
    writer = CsvResultWriter(
        source,
        output,
        AutoSavePolicy(1, 60.0),
    )

    with patch(
        "app.batch.result_writer.perf_counter",
        side_effect=[10.0, 10.4],
    ):
        assert writer.write(done_job())

    metrics = writer.autosave_metrics
    assert metrics.saves_completed == 1
    assert metrics.rows_saved == 1
    assert metrics.last_rows_saved == 1
    assert metrics.total_save_seconds == pytest.approx(0.4)
    assert metrics.average_save_seconds == pytest.approx(0.4)
    assert metrics.maximum_save_seconds == pytest.approx(0.4)
    writer.close()


def test_factory_accepts_explicit_output_path(tmp_path: Path) -> None:
    source = tmp_path / "routes.csv"
    alternate = tmp_path / "alternate.csv"
    source.write_text("Origin,Destination,Distance\nA,B,\n", encoding="utf-8")

    writer = ResultWriterFactory().create(
        source,
        "Routes",
        output_path=alternate,
    )
    assert writer.output_path == alternate
    writer.close()


def test_excel_writer_wraps_save_permission_error(tmp_path: Path) -> None:

    source = tmp_path / "routes.xlsx"
    output = tmp_path / "routes.result.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Routes"
    sheet.append(["Origin", "Destination", "Distance"])
    sheet.append(["A", "B", None])
    workbook.save(source)
    workbook.close()

    writer = ExcelResultWriter(source, "Routes", output)
    writer.write(done_job())
    with patch.object(writer._workbook, "save", side_effect=PermissionError("locked")):
        with pytest.raises(OutputWriteError, match="Unable to save"):
            writer.flush()
    writer._dirty = False
    writer.close()


def test_csv_writer_wraps_open_error_and_cleans_temp(tmp_path: Path) -> None:
    from app.batch.file_access import OutputWriteError

    source = tmp_path / "routes.csv"
    output = tmp_path / "routes.result.csv"
    source.write_text("Origin,Destination,Distance\nA,B,\n", encoding="utf-8")
    writer = CsvResultWriter(source, output)
    writer.write(done_job())

    with patch("pathlib.Path.open", side_effect=PermissionError("locked")):
        with pytest.raises(OutputWriteError, match="Unable to save"):
            writer.flush()
    writer._dirty = False
    writer.close()


def test_excel_writer_reraises_output_write_error(
    tmp_path: Path,
) -> None:

    source = tmp_path / "routes.xlsx"
    output = tmp_path / "routes.result.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Routes"
    sheet.append(["Origin", "Destination", "Distance"])
    sheet.append(["A", "B", None])
    workbook.save(source)
    workbook.close()

    writer = ExcelResultWriter(source, "Routes", output)

    error = OutputWriteError(
        output,
        "replace",
        "locked",
    )

    with patch(
        "app.batch.result_writer.AtomicOutputFile.replace",
        side_effect=error,
    ):
        with pytest.raises(OutputWriteError) as raised:
            writer._save()

    assert raised.value is error
    writer.close()


def test_excel_writer_cleans_up_and_reraises_unknown_error(
    tmp_path: Path,
) -> None:
    source = tmp_path / "routes.xlsx"
    output = tmp_path / "routes.result.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Routes"
    sheet.append(["Origin", "Destination", "Distance"])
    workbook.save(source)
    workbook.close()

    writer = ExcelResultWriter(source, "Routes", output)

    with (
        patch.object(
            writer._workbook,
            "save",
            side_effect=RuntimeError("unexpected"),
        ),
        patch(
            "app.batch.result_writer.AtomicOutputFile.cleanup",
        ) as cleanup,
    ):
        with pytest.raises(RuntimeError, match="unexpected"):
            writer._save()

    cleanup.assert_called_once_with()
    writer.close()


def test_csv_writer_reraises_output_write_error(
    tmp_path: Path,
) -> None:
    source = tmp_path / "routes.csv"
    source.write_text(
        "Origin,Destination,Distance\n" "A,B,\n",
        encoding="utf-8",
    )

    output = tmp_path / "routes.result.csv"
    writer = ResultWriterFactory().create(
        source,
        "",
        output_path=output,
    )

    expected_error = OutputWriteError(
        output,
        "replace",
        "file is locked",
    )

    temporary = MagicMock()
    temporary.open.return_value.__enter__.return_value = StringIO()

    with patch(
        "app.batch.result_writer.AtomicOutputFile",
    ) as atomic_type:
        atomic = atomic_type.return_value
        atomic.create.return_value = temporary
        atomic.replace.side_effect = expected_error

        with pytest.raises(OutputWriteError) as raised:
            writer._save()  # type: ignore[attr-defined]

    assert raised.value is expected_error
    atomic.cleanup.assert_not_called()

    writer.close()


def test_csv_writer_cleans_up_and_reraises_unexpected_error(
    tmp_path: Path,
) -> None:
    source = tmp_path / "routes.csv"
    source.write_text(
        "Origin,Destination,Distance\n" "A,B,\n",
        encoding="utf-8",
    )

    output = tmp_path / "routes.result.csv"
    writer = ResultWriterFactory().create(
        source,
        "",
        output_path=output,
    )

    temporary = MagicMock()
    temporary.open.side_effect = RuntimeError("unexpected csv failure")

    with patch(
        "app.batch.result_writer.AtomicOutputFile",
    ) as atomic_type:
        atomic = atomic_type.return_value
        atomic.create.return_value = temporary

        with pytest.raises(
            RuntimeError,
            match="unexpected csv failure",
        ):
            writer._save()  # type: ignore[attr-defined]

    atomic.cleanup.assert_called_once_with()
    atomic.replace.assert_not_called()

    writer.close()


def test_factory_never_writes_back_to_the_opened_source_path(tmp_path: Path) -> None:
    source = tmp_path / "routes.result.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Routes"
    sheet.append(["Origin", "Destination", "Distance"])
    sheet.append(["A", "B", None])
    workbook.save(source)
    workbook.close()

    writer = ResultWriterFactory().create(source, "Routes")

    assert writer.output_path == tmp_path / "routes.result.result.xlsx"
    assert writer.output_path != source
    writer.write(done_job())
    writer.close()
    assert writer.output_path.is_file()
    assert source.is_file()


def test_factory_redirects_explicit_output_equal_to_source(tmp_path: Path) -> None:
    source = tmp_path / "routes.csv"
    source.write_text("Origin,Destination,Distance\nA,B,\n", encoding="utf-8")

    writer = ResultWriterFactory().create(
        source,
        "Routes",
        output_path=source,
    )

    assert writer.output_path == tmp_path / "routes.result.csv"
    writer.close()


def test_excel_writer_writes_route_duration_to_mapped_column(tmp_path: Path) -> None:
    source = tmp_path / "routes.xlsx"
    output = tmp_path / "routes.result.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Routes"
    sheet.append(["Origin", "Destination", "Distance", "Travel time"])
    sheet.append(["A", "B", None, None])
    workbook.save(source)
    workbook.close()

    writer = ExcelResultWriter(source, "Routes", output, AutoSavePolicy(1, 60.0))
    job = RouteJob(
        2,
        "A",
        "B",
        "Distance",
        result_duration_column="Travel time",
        status=RouteJobStatus.DONE,
        result_distance_km=8.6,
        result_duration_minutes=20,
        result_duration_text="20 min",
    )

    assert writer.write(job)
    writer.close()

    saved = load_workbook(output, data_only=True)
    assert saved["Routes"]["C2"].value == 8.6
    assert saved["Routes"]["D2"].value == 20
    saved.close()


def test_csv_writer_writes_route_duration_to_mapped_column(tmp_path: Path) -> None:
    source = tmp_path / "routes.csv"
    output = tmp_path / "routes.result.csv"
    source.write_text(
        "Origin,Destination,Distance,Travel time\nA,B,,\n",
        encoding="utf-8",
    )
    writer = CsvResultWriter(source, output, AutoSavePolicy(1, 60.0))
    job = RouteJob(
        2,
        "A",
        "B",
        "Distance",
        result_duration_column="Travel time",
        status=RouteJobStatus.DONE,
        result_distance_km=8.6,
        result_duration_minutes=20,
    )

    assert writer.write(job)
    writer.close()

    with output.open("r", encoding="utf-8-sig", newline="") as stream:
        rows = list(csv.reader(stream))
    assert rows[1][2:] == ["8.6", "20"]


def test_writers_validate_result_duration_column(tmp_path: Path) -> None:
    source = tmp_path / "routes.xlsx"
    output = tmp_path / "routes.result.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Routes"
    sheet.append(["Origin", "Destination", "Distance"])
    sheet.append(["A", "B", None])
    workbook.save(source)
    workbook.close()

    excel_writer = ExcelResultWriter(source, "Routes", output)
    job = RouteJob(
        2,
        "A",
        "B",
        "Distance",
        result_duration_column="Missing duration",
        status=RouteJobStatus.DONE,
        result_distance_km=1.0,
        result_duration_minutes=5,
    )
    with pytest.raises(ValueError, match="Result duration column not found"):
        excel_writer.write(job)
    excel_writer.close()

    csv_source = tmp_path / "routes.csv"
    csv_output = tmp_path / "routes.result.csv"
    csv_source.write_text("Origin,Destination,Distance\nA,B,\n", encoding="utf-8")
    csv_writer = CsvResultWriter(csv_source, csv_output)
    with pytest.raises(ValueError, match="Result duration column not found"):
        csv_writer.write(job)
    csv_writer.close()



def test_csv_writer_extends_short_row_for_duration_column(tmp_path: Path) -> None:
    source = tmp_path / "routes.csv"
    source.write_text(
        "Origin,Destination,Distance,Travel time\nA,B\n",
        encoding="utf-8",
    )
    output = tmp_path / "routes.result.csv"
    writer = CsvResultWriter(source, output)
    job = RouteJob(
        row_index=2,
        origin="A",
        destination="B",
        result_column="Distance",
        result_duration_column="Travel time",
        status=RouteJobStatus.DONE,
        result_distance_km=8.6,
        result_duration_minutes=17,
    )

    assert writer.write(job)
    writer.flush()

    with output.open("r", encoding="utf-8-sig", newline="") as stream:
        rows = list(csv.reader(stream))
    assert rows[1] == ["A", "B", "8.6", "17"]


def test_csv_writer_rejects_invalid_row_for_duration(tmp_path: Path) -> None:
    source = tmp_path / "routes.csv"
    source.write_text(
        "Origin,Destination,Distance,Travel time\nA,B,,\n",
        encoding="utf-8",
    )
    writer = CsvResultWriter(source, tmp_path / "routes.result.csv")
    job = RouteJob(
        row_index=99,
        origin="A",
        destination="B",
        result_column="Distance",
        result_duration_column="Travel time",
        status=RouteJobStatus.DONE,
        result_distance_km=8.6,
        result_duration_minutes=17,
    )

    with pytest.raises(ValueError, match="CSV row not found: 99"):
        writer._write_duration(job, 17)
    writer.close()
