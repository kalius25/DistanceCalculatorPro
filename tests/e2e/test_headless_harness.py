from __future__ import annotations

import csv
import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from app.batch import RetryPolicy
from app.e2e import (
    E2EReportWriter,
    FakeRouteOutcome,
    HeadlessE2EHarness,
    ScriptedRouteProvider,
)
from app.enums.provider_type import ProviderType
from app.enums.travel_mode import TravelMode
from app.exceptions import ErrorCode
from app.models.route_request import RouteRequest
from app.presentation.workspace_configuration import (
    ColumnMapping,
    ProviderConfiguration,
    WorkspaceConfiguration,
)

pytestmark = pytest.mark.e2e


def configuration(*, skip_existing: bool = True) -> WorkspaceConfiguration:
    return WorkspaceConfiguration(
        ColumnMapping("Origin", "Destination", "Distance"),
        ProviderConfiguration(ProviderType.GOOGLE_MAPS_WEB, TravelMode.DRIVING),
        skip_existing_results=skip_existing,
    )


def write_csv(path: Path, rows: list[tuple[object, object, object]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.writer(stream)
        writer.writerow(("Origin", "Destination", "Distance"))
        writer.writerows(rows)


def read_distances(path: Path) -> list[str]:
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        rows = list(csv.reader(stream))
    return [row[2] for row in rows[1:]]


@pytest.mark.smoke
def test_headless_success_writes_results_and_balances_batch_lifecycle(
    tmp_path: Path,
) -> None:
    source = tmp_path / "routes.csv"
    output = tmp_path / "routes.result.csv"
    write_csv(source, [("A", "B", ""), ("C", "D", "")])
    provider = ScriptedRouteProvider(
        {
            ("A", "B"): [FakeRouteOutcome.route(12.5, 20)],
            ("C", "D"): [FakeRouteOutcome.route(7.25, 12)],
        }
    )

    report = HeadlessE2EHarness(provider).run(
        scenario="success",
        source_path=source,
        sheet_name="Routes",
        configuration=configuration(),
        output_path=output,
    )

    assert report.successful is True
    assert report.summary.successful == 2
    assert report.statuses == ("done", "done")
    assert read_distances(output) == ["12.5", "7.25"]
    assert provider.requests == 2
    assert provider.batches_started == provider.batches_finished == 1


def test_headless_transient_failure_is_retried(tmp_path: Path) -> None:
    source = tmp_path / "routes.csv"
    output = tmp_path / "routes.result.csv"
    write_csv(source, [("A", "B", "")])
    provider = ScriptedRouteProvider(
        {
            ("A", "B"): [
                FakeRouteOutcome.failure("network timeout", ErrorCode.NETWORK_ERROR),
                FakeRouteOutcome.route(9.0),
            ]
        }
    )
    harness = HeadlessE2EHarness(
        provider,
        retry_policy=RetryPolicy(
            max_attempts=2,
            initial_delay_seconds=0,
            max_delay_seconds=0,
        ),
    )

    report = harness.run(
        scenario="retry",
        source_path=source,
        sheet_name="Routes",
        configuration=configuration(),
        output_path=output,
    )

    assert report.summary.successful == 1
    assert report.summary.retry_count == 1
    assert provider.requests == 2
    assert read_distances(output) == ["9.0"]


def test_headless_terminal_failure_is_written(tmp_path: Path) -> None:
    source = tmp_path / "routes.csv"
    output = tmp_path / "routes.result.csv"
    write_csv(source, [("A", "B", "")])
    provider = ScriptedRouteProvider(
        {
            ("A", "B"): [
                FakeRouteOutcome.failure("parser mismatch", ErrorCode.PARSER_ERROR)
            ]
        }
    )

    report = HeadlessE2EHarness(provider).run(
        scenario="failure",
        source_path=source,
        sheet_name="Routes",
        configuration=configuration(),
        output_path=output,
    )

    assert report.successful is False
    assert report.summary.failed == 1
    assert report.statuses == ("failed",)
    assert read_distances(output) == ["ERROR: parser mismatch"]


def test_headless_resume_skips_existing_results(tmp_path: Path) -> None:
    source = tmp_path / "routes.csv"
    output = tmp_path / "routes.result.csv"
    write_csv(source, [("A", "B", ""), ("C", "D", "")])
    write_csv(output, [("A", "B", "4.5"), ("C", "D", "")])
    provider = ScriptedRouteProvider({("C", "D"): [FakeRouteOutcome.route(8.0)]})

    report = HeadlessE2EHarness(provider).run(
        scenario="resume",
        source_path=source,
        sheet_name="Routes",
        configuration=configuration(skip_existing=True),
        output_path=output,
        resume_from_output=True,
    )

    assert report.summary.resumed == 1
    assert report.summary.successful == 2
    assert provider.requests == 1
    assert read_distances(output) == ["4.5", "8.0"]


def test_headless_stop_retains_incremental_output(tmp_path: Path) -> None:
    source = tmp_path / "routes.csv"
    output = tmp_path / "routes.result.csv"
    write_csv(source, [("A", "B", ""), ("C", "D", ""), ("E", "F", "")])
    provider = ScriptedRouteProvider(default_distance_km=3.0)

    report = HeadlessE2EHarness(provider).run(
        scenario="stop",
        source_path=source,
        sheet_name="Routes",
        configuration=configuration(),
        output_path=output,
        stop_after=1,
    )

    assert report.summary.stopped is True
    assert report.summary.successful == 1
    assert report.statuses == ("done", "pending", "pending")
    assert read_distances(output) == ["3.0", "", ""]
    assert provider.batches_started == provider.batches_finished == 1


def test_headless_xlsx_writes_results(tmp_path: Path) -> None:
    suffix = ".xlsx"
    source = tmp_path / "routes.xlsx"
    output = tmp_path / "routes.result.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Routes"
    sheet.append(("Origin", "Destination", "Distance"))
    sheet.append(("A", "B", None))
    workbook.save(source)
    workbook.close()

    report = HeadlessE2EHarness(ScriptedRouteProvider(default_distance_km=6.75)).run(
        scenario=f"excel-{suffix[1:]}",
        source_path=source,
        sheet_name="Routes",
        configuration=configuration(),
        output_path=output,
    )

    result_workbook = load_workbook(
        output,
        data_only=True,
        keep_vba=False,
    )
    try:
        assert result_workbook["Routes"]["C2"].value == 6.75
    finally:
        result_workbook.close()
    assert report.summary.successful == 1


@pytest.mark.smoke
def test_headless_report_is_serialized(tmp_path: Path) -> None:
    source = tmp_path / "routes.csv"
    output = tmp_path / "routes.result.csv"
    reports = tmp_path / "reports"
    write_csv(source, [("A", "B", "")])
    provider = ScriptedRouteProvider(default_distance_km=2.0)

    report = HeadlessE2EHarness(
        provider,
        report_writer=E2EReportWriter(reports),
    ).run(
        scenario="smoke",
        source_path=source,
        sheet_name="Routes",
        configuration=configuration(),
        output_path=output,
        write_report=True,
    )

    assert report.report_file is not None
    payload = json.loads(Path(report.report_file).read_text(encoding="utf-8"))
    assert payload["scenario"] == "smoke"
    assert payload["successful"] is True
    assert payload["summary"]["successful"] == 1


def test_scripted_provider_can_raise_scripted_exception() -> None:
    provider = ScriptedRouteProvider(
        {
            ("A", "B"): [
                FakeRouteOutcome(
                    False,
                    exception=RuntimeError("scripted crash"),
                )
            ]
        }
    )
    with pytest.raises(RuntimeError, match="scripted crash"):
        provider.calculate(RouteRequest("A", "B"))
