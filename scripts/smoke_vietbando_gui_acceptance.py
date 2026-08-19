"""Real GUI acceptance smoke for VietBanDo."""

from __future__ import annotations

import argparse
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from PySide6.QtCore import QEventLoop, QTimer
from PySide6.QtWidgets import QMessageBox

from app.enums.travel_mode import TravelMode
from app.models.route_result import RouteResult
from app.presentation.app import create_application

_ORIGIN = "10.113922624804262,105.69436247381175"
_DESTINATION = "10.892645,105.041044"


def _create_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Routes"
    sheet.append(("From", "To", "Result distance", "Result duration"))
    sheet.append((_ORIGIN, _DESTINATION, "", ""))
    workbook.save(path)
    workbook.close()


def _read_result(path: Path) -> tuple[object, object]:
    workbook = load_workbook(path, data_only=True)
    try:
        sheet = workbook["Routes"]
        return sheet["C2"].value, sheet["D2"].value
    finally:
        workbook.close()


def _distance_is_success(value: object) -> bool:
    """Return whether the workbook distance cell contains a real result."""
    if value in (None, ""):
        return False
    if isinstance(value, str) and value.lstrip().startswith("ERROR:"):
        return False
    return True


def _print_route_diagnostics(results: object) -> None:
    """Print the original provider/engine error behind a failed row."""
    if not isinstance(results, list):
        return

    for index, result in enumerate(results, start=1):
        if not isinstance(result, RouteResult):
            continue

        print(f"RouteResult[{index}] success: {result.success}")
        print(f"RouteResult[{index}] provider: {result.provider}")
        if result.error_code is not None:
            print(f"RouteResult[{index}] error code: " f"{result.error_code.value}")
        if result.error:
            print(f"RouteResult[{index}] error: {result.error}")
        if result.context:
            print(f"RouteResult[{index}] context: {result.context}")

        exception = result.exception
        if exception is None:
            continue

        print(
            f"RouteResult[{index}] exception: "
            f"{type(exception).__name__}: {exception}"
        )
        cause = getattr(exception, "cause", None)
        if cause is not None:
            print(f"RouteResult[{index}] cause: " f"{type(cause).__name__}: {cause}")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Run VietBanDo GUI acceptance.")
    parser.add_argument(
        "--mode",
        choices=("driving", "truck", "walking"),
        default="driving",
    )
    parser.add_argument("--timeout-ms", type=int, default=90_000)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("artifacts/vietbando-gui-acceptance"),
    )
    args = parser.parse_args(argv)

    mode = TravelMode(args.mode)
    args.output.mkdir(parents=True, exist_ok=True)
    source = args.output / f"vietbando_{mode.value}.xlsx"
    result = source.with_name(f"{source.stem}.result.xlsx")
    _create_workbook(source)
    if result.exists():
        result.unlink()

    application, window, exception_handler, splash = create_application()
    splash.finish(window)
    completed = False
    failed = False
    error = ""
    completed_results: object = None

    def on_completed(results: object) -> None:
        nonlocal completed, completed_results
        completed = True
        completed_results = results
        loop.quit()

    def on_failed(message: str) -> None:
        nonlocal failed, error
        failed = True
        error = message
        loop.quit()

    loop = QEventLoop()
    timer = QTimer()
    timer.setSingleShot(True)
    timer.timeout.connect(loop.quit)

    try:
        with patch.object(
            QMessageBox,
            "question",
            return_value=QMessageBox.StandardButton.No,
        ):
            window._select_workbook(str(source))
            window._home_page._provider_selector.setCurrentText("VietBanDo")
            label = {
                TravelMode.DRIVING: "Driving",
                TravelMode.TRUCK: "Truck",
                TravelMode.WALKING: "Walking",
            }[mode]
            window._home_page._travel_mode_selector.setCurrentText(label)
            application.processEvents()

            window.calculation_completed.connect(on_completed)
            window.calculation_failed.connect(on_failed)
            timer.start(args.timeout_ms)
            window._action_start.trigger()
            application.processEvents()
            loop.exec()
            timer.stop()

        distance: object = None
        duration: object = None
        if completed and result.is_file():
            distance, duration = _read_result(result)

        passed = (
            completed
            and not failed
            and result.is_file()
            and _distance_is_success(distance)
            and duration in (None, "")
        )

        print(f"GUI acceptance: {'PASS' if passed else 'FAIL'}")
        print(f"Completed: {completed}")
        print(f"Failed: {failed}")
        print(f"Result distance: {distance}")
        print(f"Result duration: {duration!r}")
        if error:
            print(f"Error: {error}")
        if not passed:
            _print_route_diagnostics(completed_results)
        return 0 if passed else 1
    finally:
        window.shutdown()
        exception_handler.restore()
        window.close()
        application.processEvents()
        application.quit()


if __name__ == "__main__":
    raise SystemExit(main())
