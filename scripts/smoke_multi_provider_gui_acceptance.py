"""Run the v1.3 RC multi-provider GUI regression matrix."""

from __future__ import annotations

import argparse
import json
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from PySide6.QtCore import QEventLoop, QTimer
from PySide6.QtWidgets import QMessageBox

from app.presentation.app import create_application
from app.release.provider_regression_matrix import (
    PROVIDER_REGRESSION_CASES,
    ProviderRegressionCase,
    distance_is_success,
    duration_matches_capability,
)

_ORIGIN = "10.113922624804262,105.69436247381175"
_DESTINATION = "10.892645,105.041044"

_MODE_LABEL = {
    "driving": "Driving",
    "truck": "Truck",
    "walking": "Walking",
}


@dataclass(slots=True)
class RegressionResult:
    provider: str
    travel_mode: str
    duration_required: bool
    passed: bool
    completed: bool
    failed: bool
    output_exists: bool
    worker_idle: bool
    distance: object
    duration: object
    error: str


def _create_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Routes"
    sheet.append(("From", "To", "Result distance", "Result duration"))
    sheet.append((_ORIGIN, _DESTINATION, "", ""))
    workbook.save(path)
    workbook.close()


def _read_output(path: Path) -> tuple[object, object]:
    workbook = load_workbook(path, data_only=True)
    try:
        sheet = workbook["Routes"]
        return sheet["C2"].value, sheet["D2"].value
    finally:
        workbook.close()


def _wait_for_coordinator_idle(
    application: object,
    window: object,
    timeout_ms: int = 10_000,
) -> bool:
    """Wait until the previous worker QThread is fully destroyed."""
    if not window._execution_coordinator.is_running:
        return True

    loop = QEventLoop()
    timer = QTimer()
    poll = QTimer()

    timer.setSingleShot(True)
    timer.timeout.connect(loop.quit)

    def check_idle() -> None:
        application.processEvents()
        if not window._execution_coordinator.is_running:
            loop.quit()

    poll.timeout.connect(check_idle)
    poll.start(25)
    timer.start(timeout_ms)
    loop.exec()
    poll.stop()
    timer.stop()

    application.processEvents()
    return not window._execution_coordinator.is_running


def _run_case(
    *,
    application: object,
    window: object,
    case: ProviderRegressionCase,
    artifact_dir: Path,
    timeout_ms: int,
) -> RegressionResult:
    stem = case.provider.name.lower() + "_" + case.travel_mode.value
    source = artifact_dir / f"{stem}.xlsx"
    output = artifact_dir / f"{stem}.result.xlsx"
    _create_workbook(source)
    if output.exists():
        output.unlink()

    completed = False
    failed = False
    error = ""
    loop = QEventLoop()

    def on_completed(_results: object) -> None:
        nonlocal completed
        completed = True
        loop.quit()

    def on_failed(message: str) -> None:
        nonlocal failed, error
        failed = True
        error = message
        loop.quit()

    window.calculation_completed.connect(on_completed)
    window.calculation_failed.connect(on_failed)

    timer = QTimer()
    timer.setSingleShot(True)
    timer.timeout.connect(loop.quit)

    if not _wait_for_coordinator_idle(application, window):
        return RegressionResult(
            provider=case.provider.value,
            travel_mode=case.travel_mode.value,
            duration_required=case.duration_required,
            passed=False,
            completed=False,
            failed=True,
            output_exists=False,
            worker_idle=False,
            distance=None,
            duration=None,
            error="Previous calculation worker did not shut down.",
        )

    with patch.object(
        QMessageBox,
        "question",
        return_value=QMessageBox.StandardButton.No,
    ):
        window._select_workbook(str(source))
        window._home_page._provider_selector.setCurrentText(case.provider.value)
        window._home_page._travel_mode_selector.setCurrentText(
            _MODE_LABEL[case.travel_mode.value]
        )
        application.processEvents()

        timer.start(timeout_ms)
        window._action_start.trigger()
        application.processEvents()

        if not window._execution_coordinator.is_running:
            timer.stop()
            return RegressionResult(
                provider=case.provider.value,
                travel_mode=case.travel_mode.value,
                duration_required=case.duration_required,
                passed=False,
                completed=False,
                failed=True,
                output_exists=output.is_file(),
                worker_idle=False,
                distance=None,
                duration=None,
                error="Calculation worker did not start.",
            )

        loop.exec()
        timer.stop()

    worker_idle = _wait_for_coordinator_idle(
        application,
        window,
    )
    if not worker_idle and not error:
        error = "Calculation completed but worker thread did not shut down."

    distance: object = None
    duration: object = None
    if completed and output.is_file():
        distance, duration = _read_output(output)

    passed = (
        worker_idle
        and completed
        and not failed
        and output.is_file()
        and distance_is_success(distance)
        and duration_matches_capability(
            duration,
            required=case.duration_required,
        )
    )

    try:
        window.calculation_completed.disconnect(on_completed)
        window.calculation_failed.disconnect(on_failed)
    except RuntimeError:
        pass

    return RegressionResult(
        provider=case.provider.value,
        travel_mode=case.travel_mode.value,
        duration_required=case.duration_required,
        passed=passed,
        completed=completed,
        failed=failed,
        output_exists=output.is_file(),
        worker_idle=worker_idle,
        distance=distance,
        duration=duration,
        error=error,
    )


def _print_result(result: RegressionResult) -> None:
    print()
    print(f"== {result.provider} / {result.travel_mode} ==")
    print(f"Regression: {'PASS' if result.passed else 'FAIL'}")
    print(f"Completed: {result.completed}")
    print(f"Failed: {result.failed}")
    print(f"Output exists: {result.output_exists}")
    print(f"Worker idle: {result.worker_idle}")
    print(f"Result distance: {result.distance}")
    print(f"Result duration: {result.duration!r}")
    print(f"Duration required: {result.duration_required}")
    if result.error:
        print(f"Error: {result.error}")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Run all v1.3 production provider GUI regressions."
    )
    parser.add_argument(
        "--timeout-ms",
        type=int,
        default=90_000,
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("artifacts/multi-provider-gui-regression"),
    )
    args = parser.parse_args(argv)

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    artifact_dir = args.output / timestamp
    artifact_dir.mkdir(parents=True, exist_ok=True)

    application, window, exception_handler, splash = create_application()
    splash.finish(window)

    try:
        results = [
            _run_case(
                application=application,
                window=window,
                case=case,
                artifact_dir=artifact_dir,
                timeout_ms=args.timeout_ms,
            )
            for case in PROVIDER_REGRESSION_CASES
        ]
    finally:
        window.shutdown()
        exception_handler.restore()
        window.close()
        application.processEvents()
        application.quit()

    for result in results:
        _print_result(result)

    report_path = artifact_dir / "report.json"
    report_path.write_text(
        json.dumps(
            [asdict(result) for result in results],
            ensure_ascii=False,
            indent=2,
            default=str,
        ),
        encoding="utf-8",
    )

    passed_count = sum(result.passed for result in results)
    print()
    print(f"Regression summary: {passed_count}/{len(results)} PASS")
    print(f"Report: {report_path}")

    return 0 if passed_count == len(results) else 1


if __name__ == "__main__":
    raise SystemExit(main())
