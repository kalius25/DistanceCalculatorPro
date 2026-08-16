"""Real GUI acceptance smoke for Bing Maps and OpenStreetMap."""

from __future__ import annotations

import argparse
import json
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from PySide6.QtCore import QEventLoop, QTimer
from PySide6.QtWidgets import QMessageBox

from app.enums.provider_type import ProviderType
from app.enums.travel_mode import TravelMode
from app.presentation.app import create_application

_ORIGIN = "10.113922624804262,105.69436247381175"
_DESTINATION = "10.892645,105.041044"


@dataclass(slots=True)
class GuiAcceptanceResult:
    provider: str
    travel_mode: str
    passed: bool
    workspace_ready: bool
    calculation_completed: bool
    calculation_failed: bool
    output_exists: bool
    distance_value: object
    duration_value: object
    final_status: str
    error: str


def _create_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Routes"
    sheet.append(("From", "To", "Result distance", "Result duration"))
    sheet.append((_ORIGIN, _DESTINATION, "", ""))
    workbook.save(path)
    workbook.close()


def _read_outputs(path: Path) -> tuple[object, object]:
    workbook = load_workbook(path, data_only=True)
    try:
        sheet = workbook["Routes"]
        return sheet["C2"].value, sheet["D2"].value
    finally:
        workbook.close()


def _wait_for_completion(
    application: object,
    window: object,
    timeout_ms: int,
) -> tuple[bool, bool, str]:
    loop = QEventLoop()
    completed = False
    failed = False
    error = ""

    def on_completed(_results: object) -> None:
        nonlocal completed
        completed = True
        loop.quit()

    def on_failed(message: str) -> None:
        nonlocal failed, error
        failed = True
        error = message
        loop.quit()

    window.calculation_completed.connect(on_completed)
    window.calculation_failed.connect(on_failed)

    timer = QTimer()
    timer.setSingleShot(True)
    timer.timeout.connect(loop.quit)
    timer.start(timeout_ms)

    window._action_start.trigger()
    application.processEvents()
    loop.exec()

    timer.stop()
    return completed, failed, error


def _run_provider(
    *,
    application: object,
    window: object,
    provider: ProviderType,
    travel_mode: TravelMode,
    artifact_dir: Path,
    timeout_ms: int,
) -> GuiAcceptanceResult:
    source_path = artifact_dir / (
        provider.name.lower() + "_" + travel_mode.value + ".xlsx"
    )
    result_path = source_path.with_name(
        f"{source_path.stem}.result{source_path.suffix}"
    )
    _create_workbook(source_path)
    if result_path.exists():
        result_path.unlink()

    completed = False
    failed = False
    error = ""
    distance: object = None
    duration: object = None

    with patch.object(
        QMessageBox,
        "question",
        return_value=QMessageBox.StandardButton.No,
    ):
        window._select_workbook(str(source_path))
        window._home_page._provider_selector.setCurrentText(provider.value)
        mode_label = "Driving" if travel_mode is TravelMode.DRIVING else "Walking"
        window._home_page._travel_mode_selector.setCurrentText(mode_label)
        application.processEvents()

        workspace_ready = window._home_page.workspace_ready
        if not workspace_ready:
            error = "Workspace did not become ready."
        else:
            completed, failed, error = _wait_for_completion(
                application,
                window,
                timeout_ms,
            )

        if completed and result_path.is_file():
            distance, duration = _read_outputs(result_path)

        passed = (
            workspace_ready
            and completed
            and not failed
            and result_path.is_file()
            and distance not in (None, "")
            and duration not in (None, "")
        )

        return GuiAcceptanceResult(
            provider=provider.value,
            travel_mode=travel_mode.value,
            passed=passed,
            workspace_ready=workspace_ready,
            calculation_completed=completed,
            calculation_failed=failed,
            output_exists=result_path.is_file(),
            distance_value=distance,
            duration_value=duration,
            final_status=window._status_label.text(),
            error=error,
        )


def _print_result(result: GuiAcceptanceResult) -> None:
    print()
    print(f"== {result.provider} / {result.travel_mode} ==")
    print(f"GUI acceptance: {'PASS' if result.passed else 'FAIL'}")
    print(f"Workspace ready: {result.workspace_ready}")
    print(f"Completed signal: {result.calculation_completed}")
    print(f"Failed signal: {result.calculation_failed}")
    print(f"Output exists: {result.output_exists}")
    print(f"Result distance: {result.distance_value}")
    print(f"Result duration: {result.duration_value}")
    print(f"Final status: {result.final_status}")
    if result.error:
        print(f"Error: {result.error}")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Run real GUI acceptance for Bing Maps/OpenStreetMap."
    )
    parser.add_argument(
        "--provider",
        choices=("all", "bing", "osm"),
        default="all",
    )
    parser.add_argument(
        "--mode",
        choices=("driving", "walking"),
        default="driving",
    )
    parser.add_argument(
        "--timeout-ms",
        type=int,
        default=90_000,
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("artifacts/gui-provider-acceptance"),
    )
    args = parser.parse_args(argv)

    providers: list[ProviderType] = []
    if args.provider in ("all", "bing"):
        providers.append(ProviderType.BING_MAPS_WEB)
    if args.provider in ("all", "osm"):
        providers.append(ProviderType.OPENSTREETMAP_WEB)

    travel_mode = TravelMode(args.mode)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    artifact_dir = args.output / timestamp
    artifact_dir.mkdir(parents=True, exist_ok=True)

    application, window, exception_handler, splash = create_application()
    splash.finish(window)

    try:
        results = [
            _run_provider(
                application=application,
                window=window,
                provider=provider,
                travel_mode=travel_mode,
                artifact_dir=artifact_dir,
                timeout_ms=args.timeout_ms,
            )
            for provider in providers
        ]
    finally:
        window.shutdown()
        exception_handler.restore()
        window.close()
        application.processEvents()
        application.quit()

    for result in results:
        _print_result(result)

    report_path = artifact_dir / "report.json"
    report_path.write_text(
        json.dumps(
            [asdict(result) for result in results],
            ensure_ascii=False,
            indent=2,
            default=str,
        ),
        encoding="utf-8",
    )
    print()
    print(f"Report: {report_path}")

    return 0 if all(result.passed for result in results) else 1


if __name__ == "__main__":
    raise SystemExit(main())
