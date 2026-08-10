"""Incremental workbook result writers."""

from __future__ import annotations

import csv
from abc import ABC, abstractmethod
from io import BytesIO
from pathlib import Path
from time import perf_counter

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.logging import LoggingManager

from .autosave_metrics import AutosaveMetrics, AutosaveSnapshot
from .autosave_policy import AutoSavePolicy
from .file_access import AtomicOutputFile, OutputWriteError, ensure_output_writable
from .models import RouteJob, RouteJobStatus
from .output_path_policy import OutputPathPolicy

logger = LoggingManager.get_logger(__name__)


class BaseResultWriter(ABC):
    """Write route-job outcomes and persist them incrementally."""

    def __init__(
        self,
        output_path: Path,
        autosave_policy: AutoSavePolicy | None = None,
    ) -> None:
        self.output_path = output_path
        self._autosave_policy = autosave_policy or AutoSavePolicy()
        self._dirty = False
        self._closed = False
        self._autosave_metrics = AutosaveMetrics()

    @property
    def dirty(self) -> bool:
        return self._dirty

    @property
    def closed(self) -> bool:
        return self._closed

    @property
    def autosave_metrics(self) -> AutosaveSnapshot:
        return self._autosave_metrics.snapshot

    def write(self, job: RouteJob) -> bool:
        value = self._value_for(job)
        if value is None:
            return False
        self._write_value(job, value)
        if (
            job.status is RouteJobStatus.DONE
            and job.result_duration_column
            and job.result_duration_minutes is not None
        ):
            self._write_duration(job, job.result_duration_minutes)
        self._dirty = True
        self._autosave_policy.record_write()
        if self._autosave_policy.should_save():
            self.flush()
        return True

    def flush(self) -> bool:
        if self._closed or not self._dirty:
            return False
        logger.debug(
            "RESULT_FLUSH_BEGIN",
            extra={
                "event": "RESULT_FLUSH_BEGIN",
                "writer_type": type(self).__name__,
                "output_path": str(self.output_path),
                "output_exists": self.output_path.exists(),
                "dirty_rows": self._autosave_policy.dirty_rows,
            },
        )
        ensure_output_writable(self.output_path)
        dirty_rows = self._autosave_policy.dirty_rows
        started_at = perf_counter()
        self._save()
        elapsed = perf_counter() - started_at
        self._autosave_metrics.record(dirty_rows, elapsed)
        self._dirty = False
        self._autosave_policy.mark_saved()
        logger.debug(
            "RESULT_FLUSH_OK",
            extra={
                "event": "RESULT_FLUSH_OK",
                "writer_type": type(self).__name__,
                "output_path": str(self.output_path),
                "elapsed_seconds": elapsed,
                "output_exists": self.output_path.is_file(),
                "output_size_bytes": self.output_path.stat().st_size,
            },
        )
        return True

    def close(self) -> None:
        if self._closed:
            return
        self.flush()
        self._close()
        self._closed = True

    def __enter__(self) -> BaseResultWriter:
        return self

    def __exit__(self, *_args: object) -> None:
        self.close()

    @staticmethod
    def _value_for(job: RouteJob) -> float | str | None:
        if job.status is RouteJobStatus.DONE:
            return job.result_distance_km
        if job.status in {RouteJobStatus.FAILED, RouteJobStatus.INVALID}:
            return f"ERROR: {job.validation_error or 'Unknown error.'}"
        return None

    @abstractmethod
    def _write_value(self, job: RouteJob, value: float | str) -> None:
        """Write one value into the in-memory document."""

    @abstractmethod
    def _write_duration(self, job: RouteJob, value: int | str) -> None:
        """Write route duration into the configured duration column."""

    @abstractmethod
    def _save(self) -> None:
        """Persist the in-memory document to output_path."""

    def _close(self) -> None:
        """Release resources after the final flush."""


class ExcelResultWriter(BaseResultWriter):
    """Keep one openpyxl workbook alive for a complete batch."""

    def __init__(
        self,
        source_path: str | Path,
        sheet_name: str,
        output_path: Path,
        autosave_policy: AutoSavePolicy | None = None,
    ) -> None:
        source = Path(source_path)
        # Load from an in-memory snapshot instead of giving openpyxl the live
        # source path.  This is especially important when resuming directly
        # from an existing *.result.xlsx on Windows: some openpyxl/ZipFile
        # combinations can otherwise retain a file handle long enough to make
        # the later atomic os.replace(temp, output) fail with WinError 5.
        logger.debug(
            "EXCEL_WRITER_OPEN_BEGIN",
            extra={
                "event": "EXCEL_WRITER_OPEN_BEGIN",
                "source_path": str(source),
                "output_path": str(output_path),
                "paths_are_distinct": source != output_path,
            },
        )
        source_bytes = source.read_bytes()
        source_snapshot = BytesIO(source_bytes)
        logger.debug(
            "EXCEL_WRITER_SOURCE_SNAPSHOT_OK",
            extra={
                "event": "EXCEL_WRITER_SOURCE_SNAPSHOT_OK",
                "source_path": str(source),
                "source_size_bytes": len(source_bytes),
            },
        )
        self._workbook: Workbook = load_workbook(
            source_snapshot,
            data_only=False,
            keep_vba=source.suffix.casefold() == ".xlsm",
        )
        if sheet_name not in self._workbook.sheetnames:
            self._workbook.close()
            raise ValueError(f"Worksheet not found: {sheet_name}")
        self._worksheet: Worksheet = self._workbook[sheet_name]
        self._columns = self._header_columns(self._worksheet)
        super().__init__(output_path, autosave_policy)
        logger.info(
            "EXCEL_WRITER_OPEN_OK",
            extra={
                "event": "EXCEL_WRITER_OPEN_OK",
                "source_path": str(source),
                "output_path": str(output_path),
                "sheet_name": sheet_name,
                "column_count": len(self._columns),
            },
        )

    def _write_value(self, job: RouteJob, value: float | str) -> None:
        column = self._columns.get(job.result_column)
        if column is None:
            raise ValueError(f"Result column not found: {job.result_column}")
        self._worksheet.cell(row=job.row_index, column=column, value=value)

    def _write_duration(self, job: RouteJob, value: int | str) -> None:
        column = self._columns.get(job.result_duration_column)
        if column is None:
            raise ValueError(
                f"Result duration column not found: {job.result_duration_column}"
            )
        self._worksheet.cell(row=job.row_index, column=column, value=value)

    def _save(self) -> None:
        logger.debug(
            "EXCEL_SAVE_BEGIN",
            extra={
                "event": "EXCEL_SAVE_BEGIN",
                "output_path": str(self.output_path),
                "output_exists": self.output_path.exists(),
            },
        )
        atomic = AtomicOutputFile(self.output_path)
        temporary = atomic.create(suffix=self.output_path.suffix)
        try:
            logger.debug(
                "EXCEL_SAVE_TEMP_WRITE_BEGIN",
                extra={
                    "event": "EXCEL_SAVE_TEMP_WRITE_BEGIN",
                    "output_path": str(self.output_path),
                    "temp_path": str(temporary),
                },
            )
            self._workbook.save(temporary)
            logger.debug(
                "EXCEL_SAVE_TEMP_WRITE_OK",
                extra={
                    "event": "EXCEL_SAVE_TEMP_WRITE_OK",
                    "output_path": str(self.output_path),
                    "temp_path": str(temporary),
                    "temp_exists": temporary.is_file(),
                    "temp_size_bytes": temporary.stat().st_size,
                },
            )
            atomic.replace()
            logger.debug(
                "EXCEL_SAVE_OK",
                extra={
                    "event": "EXCEL_SAVE_OK",
                    "output_path": str(self.output_path),
                },
            )
        except OutputWriteError:
            logger.exception(
                "EXCEL_SAVE_OUTPUT_WRITE_FAILED",
                extra={
                    "event": "EXCEL_SAVE_OUTPUT_WRITE_FAILED",
                    "output_path": str(self.output_path),
                    "temp_path": str(temporary),
                },
            )
            raise
        except (PermissionError, OSError) as error:
            logger.exception(
                "EXCEL_SAVE_IO_FAILED",
                extra={
                    "event": "EXCEL_SAVE_IO_FAILED",
                    "output_path": str(self.output_path),
                    "temp_path": str(temporary),
                    "error_type": type(error).__name__,
                    "error": str(error),
                },
            )
            atomic.cleanup()
            raise OutputWriteError(
                self.output_path,
                "save",
                str(error),
            ) from error
        except Exception:
            atomic.cleanup()
            raise

    def _close(self) -> None:
        self._workbook.close()

    @staticmethod
    def _header_columns(worksheet: Worksheet) -> dict[str, int]:
        return {
            str(cell.value).strip(): cell.column
            for cell in worksheet[1]
            if cell.value is not None
        }


class CsvResultWriter(BaseResultWriter):
    """Maintain CSV rows in memory and rewrite the output on autosave."""

    def __init__(
        self,
        source_path: str | Path,
        output_path: Path,
        autosave_policy: AutoSavePolicy | None = None,
    ) -> None:
        source = Path(source_path)
        with source.open("r", encoding="utf-8-sig", newline="") as stream:
            self._rows = list(csv.reader(stream))
        if not self._rows:
            raise ValueError("CSV file does not contain a header row")
        self._columns = {
            header.strip(): index for index, header in enumerate(self._rows[0])
        }
        super().__init__(output_path, autosave_policy)

    def _write_value(self, job: RouteJob, value: float | str) -> None:
        column = self._columns.get(job.result_column)
        if column is None:
            raise ValueError(f"Result column not found: {job.result_column}")
        row_index = job.row_index - 1
        if row_index < 1 or row_index >= len(self._rows):
            raise ValueError(f"CSV row not found: {job.row_index}")
        row = self._rows[row_index]
        while len(row) <= column:
            row.append("")
        row[column] = str(value)

    def _write_duration(self, job: RouteJob, value: int | str) -> None:
        column = self._columns.get(job.result_duration_column)
        if column is None:
            raise ValueError(
                f"Result duration column not found: {job.result_duration_column}"
            )
        row_index = job.row_index - 1
        if row_index < 1 or row_index >= len(self._rows):
            raise ValueError(f"CSV row not found: {job.row_index}")
        row = self._rows[row_index]
        while len(row) <= column:
            row.append("")
        row[column] = str(value)

    def _save(self) -> None:
        atomic = AtomicOutputFile(self.output_path)
        temporary = atomic.create(suffix=self.output_path.suffix)
        try:
            with temporary.open(
                "w",
                encoding="utf-8-sig",
                newline="",
            ) as stream:
                csv.writer(stream).writerows(self._rows)
            atomic.replace()
        except OutputWriteError:
            raise
        except (PermissionError, OSError) as error:
            atomic.cleanup()
            raise OutputWriteError(
                self.output_path,
                "save",
                str(error),
            ) from error
        except Exception:
            atomic.cleanup()
            raise


class ResultWriterFactory:
    """Create the correct writer for a supported workbook type."""

    def __init__(
        self,
        output_path_policy: OutputPathPolicy | None = None,
    ) -> None:
        self._output_path_policy = output_path_policy or OutputPathPolicy()

    def create(
        self,
        source_path: str | Path,
        sheet_name: str,
        autosave_policy: AutoSavePolicy | None = None,
        *,
        resume_from_output: bool = False,
        output_path: str | Path | None = None,
    ) -> BaseResultWriter:
        source = Path(source_path)
        output = (
            Path(output_path)
            if output_path is not None
            else self._output_path_policy.build(source)
        )
        if output.resolve() == source.resolve():
            output = self._output_path_policy.build(source)
        input_path = output if resume_from_output and output.exists() else source
        logger.info(
            "RESULT_WRITER_FACTORY_CREATE",
            extra={
                "event": "RESULT_WRITER_FACTORY_CREATE",
                "source_path": str(source),
                "input_path": str(input_path),
                "output_path": str(output),
                "resume_from_output": resume_from_output,
                "output_exists": output.exists(),
                "paths_are_distinct": source.resolve() != output.resolve(),
            },
        )
        suffix = source.suffix.casefold()
        if suffix in {".xlsx", ".xlsm"}:
            return ExcelResultWriter(
                input_path,
                sheet_name,
                output,
                autosave_policy,
            )
        if suffix == ".csv":
            return CsvResultWriter(input_path, output, autosave_policy)
        raise ValueError(f"Unsupported workbook type: {source.suffix}")


__all__ = [
    "BaseResultWriter",
    "CsvResultWriter",
    "ExcelResultWriter",
    "ResultWriterFactory",
]
