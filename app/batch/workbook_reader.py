"""Streaming readers for Excel and CSV batch workbooks."""

from __future__ import annotations

import csv
from collections.abc import Iterator
from pathlib import Path

from openpyxl import load_workbook

from .exceptions import BatchWorkbookError
from .models import WorkbookRow, WorkbookStream


class WorkbookReader:
    """Open supported workbook formats without loading all rows into memory."""

    SUPPORTED_SUFFIXES = frozenset({".xlsx", ".xlsm", ".csv"})

    def read(self, file_path: str | Path, sheet_name: str) -> WorkbookStream:
        path = Path(file_path)
        if not path.is_file():
            raise FileNotFoundError(f"Workbook not found: {path}")

        suffix = path.suffix.lower()
        if suffix not in self.SUPPORTED_SUFFIXES:
            raise BatchWorkbookError(f"Unsupported workbook type: {suffix}")

        if suffix == ".csv":
            return self._read_csv(path)
        return self._read_excel(path, sheet_name)

    def _read_excel(
        self,
        path: Path,
        sheet_name: str,
    ) -> WorkbookStream:
        workbook = load_workbook(
            filename=path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )

        iterator: Iterator[tuple[object, ...]] | None = None

        try:
            if sheet_name not in workbook.sheetnames:
                raise BatchWorkbookError(
                    f"Worksheet not found: {sheet_name}"
                )

            worksheet = workbook[sheet_name]
            iterator = worksheet.iter_rows(values_only=True)

            first_row = next(iterator, ())
            headers = tuple(self._text(value) for value in first_row)

        except Exception:
            if iterator is not None:
                close_iterator = getattr(iterator, "close", None)
                if callable(close_iterator):
                    close_iterator()

            workbook.close()
            raise

        def rows() -> Iterator[WorkbookRow]:
            try:
                for row_number, values in enumerate(iterator, start=2):
                    yield WorkbookRow(
                        row_number=row_number,
                        values=tuple(values),
                    )
            finally:
                close_iterator = getattr(iterator, "close", None)
                if callable(close_iterator):
                    close_iterator()

                workbook.close()

        return WorkbookStream(
            headers=headers,
            rows=rows(),
        )

    def _read_csv(self, path: Path) -> WorkbookStream:
        with path.open("r", encoding="utf-8-sig", newline="") as stream:
            reader = csv.reader(stream)
            headers = tuple(next(reader, ()))

        def rows() -> Iterator[WorkbookRow]:
            with path.open("r", encoding="utf-8-sig", newline="") as stream:
                reader = csv.reader(stream)
                next(reader, None)
                for row_number, values in enumerate(reader, start=2):
                    yield WorkbookRow(row_number, tuple(values))

        return WorkbookStream(headers, rows())

    @staticmethod
    def _text(value: object) -> str:
        return "" if value is None else str(value).strip()
