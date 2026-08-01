"""Models and input loading for calculation execution."""

from __future__ import annotations

import csv
from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from app.enums.route_preference import RoutePreference
from app.models.route_request import RouteRequest
from app.presentation.workspace_configuration import WorkspaceConfiguration


@dataclass(frozen=True, slots=True)
class CalculationJob:
    """A validated workbook calculation job."""

    file_path: str
    sheet_name: str
    configuration: WorkspaceConfiguration


Row = Sequence[object]


class CalculationJobBuilder:
    """Build route requests from the selected Excel or CSV worksheet."""

    def build_requests(self, job: CalculationJob) -> list[RouteRequest]:
        path = Path(job.file_path)
        if not path.is_file():
            raise FileNotFoundError(f"Workbook not found: {path}")

        suffix = path.suffix.lower()
        if suffix in {".xlsx", ".xlsm"}:
            headers, rows = self._read_excel(path, job.sheet_name)
        elif suffix == ".csv":
            headers, rows = self._read_csv(path)
        else:
            raise ValueError(f"Unsupported workbook type: {suffix}")

        mapping = job.configuration.column_mapping
        indexes = {header: index for index, header in enumerate(headers)}
        required = (mapping.origin_column, mapping.destination_column)
        missing = [column for column in required if column not in indexes]
        if missing:
            raise ValueError(
                "Missing mapped columns: " + ", ".join(missing)
            )

        provider = job.configuration.provider_configuration
        avoid = RoutePreference.AVOID
        automatic = RoutePreference.AUTO
        requests: list[RouteRequest] = []

        for row_number, row in enumerate(rows, start=2):
            origin = self._cell(row, indexes[mapping.origin_column])
            destination = self._cell(row, indexes[mapping.destination_column])
            if not origin or not destination:
                continue
            requests.append(
                RouteRequest(
                    origin=origin,
                    destination=destination,
                    travel_mode=provider.travel_mode,
                    toll_preference=avoid if provider.avoid_tolls else automatic,
                    highway_preference=(
                        avoid if provider.avoid_highways else automatic
                    ),
                    ferry_preference=(
                        avoid if provider.avoid_ferries else automatic
                    ),
                    metadata={
                        "row_number": row_number,
                        "result_column": mapping.result_column,
                    },
                )
            )

        return requests

    @staticmethod
    def _cell(row: Row, index: int) -> str:
        if index >= len(row):
            return ""
        value = row[index]
        return "" if value is None else str(value).strip()

    @staticmethod
    def _read_excel(
        path: Path,
        sheet_name: str,
    ) -> tuple[list[str], list[Row]]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Worksheet not found: {sheet_name}")
            worksheet = workbook[sheet_name]
            iterator = worksheet.iter_rows(values_only=True)
            first_row = next(iterator, ())
            headers = [
                "" if value is None else str(value)
                for value in first_row
            ]
            rows: list[Row] = [tuple(row) for row in iterator]
            return headers, rows
        finally:
            workbook.close()

    @staticmethod
    def _read_csv(path: Path) -> tuple[list[str], list[Row]]:
        with path.open("r", encoding="utf-8-sig", newline="") as stream:
            rows = list(csv.reader(stream))
        if not rows:
            return [], []
        data_rows: list[Row] = [list(row) for row in rows[1:]]
        return rows[0], data_rows
