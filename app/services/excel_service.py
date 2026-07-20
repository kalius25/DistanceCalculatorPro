from pathlib import Path

from openpyxl import load_workbook


class ExcelService:
    """
    Xử lý đọc dữ liệu Excel.
    """

    PREVIEW_ROWS = 100

    def __init__(self):
        self.file_path = None
        self.workbook = None

    # ==========================================================
    # Workbook
    # ==========================================================

    def open_workbook(self, file_path: str):

        self.file_path = Path(file_path)

        self.workbook = load_workbook(
            filename=self.file_path, read_only=True, data_only=True
        )

        return self.workbook.sheetnames

    def get_sheet_names(self):

        if self.workbook is None:
            return []

        return self.workbook.sheetnames

    # ==========================================================
    # Worksheet
    # ==========================================================

    def get_worksheet(self, sheet_name: str):

        if self.workbook is None:
            raise RuntimeError("Workbook chưa được mở.")

        return self.workbook[sheet_name]

    # ==========================================================
    # Header
    # ==========================================================

    def read_headers(self, sheet_name: str):

        ws = self.get_worksheet(sheet_name)

        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

        return ["" if value is None else str(value) for value in first_row]

    # ==========================================================
    # Preview
    # ==========================================================

    def read_preview(self, sheet_name: str, max_rows: int = PREVIEW_ROWS):

        ws = self.get_worksheet(sheet_name)

        rows = []

        for row in ws.iter_rows(min_row=2, max_row=max_rows + 1, values_only=True):
            rows.append(["" if cell is None else cell for cell in row])

        return rows

    # ==========================================================
    # Full Data (sẽ dùng ở Sprint sau)
    # ==========================================================

    def read_all(self, sheet_name: str):

        ws = self.get_worksheet(sheet_name)

        rows = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(["" if cell is None else cell for cell in row])

        return rows
