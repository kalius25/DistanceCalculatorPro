import csv
from pathlib import Path
from typing import Any, Protocol

from openpyxl import load_workbook

from .models import WorksheetInfo


class WorkbookReader(Protocol):
    """Port implemented by file-format-specific metadata readers."""

    def supports(self, file_path: Path) -> bool: ...

    def read_worksheets(self, file_path: Path) -> tuple[WorksheetInfo, ...]: ...


class OpenPyXLWorkbookReader:
    """Read Excel metadata and a bounded preview in read-only mode."""

    SUPPORTED_EXTENSIONS = frozenset({".xlsx", ".xlsm"})
    PREVIEW_ROW_LIMIT = 500

    def supports(self, file_path: Path) -> bool:
        return file_path.suffix.lower() in self.SUPPORTED_EXTENSIONS

    def read_worksheets(self, file_path: Path) -> tuple[WorksheetInfo, ...]:
        workbook = load_workbook(
            filename=file_path,
            read_only=True,
            data_only=True,
            keep_vba=file_path.suffix.lower() == ".xlsm",
        )
        try:
            worksheets: list[WorksheetInfo] = []
            for worksheet in workbook.worksheets:
                rows = worksheet.iter_rows(
                    min_row=1,
                    max_row=self.PREVIEW_ROW_LIMIT + 1,
                    values_only=True,
                )
                first_row = next(rows, ())
                headers = self._normalize_row(first_row)
                preview_rows = tuple(
                    self._normalize_row(row) for row in rows
                )
                worksheets.append(
                    WorksheetInfo(
                        name=worksheet.title,
                        row_count=worksheet.max_row or 0,
                        column_count=worksheet.max_column or 0,
                        headers=headers,
                        preview_rows=preview_rows,
                    )
                )
            return tuple(worksheets)
        finally:
            workbook.close()

    @staticmethod
    def _normalize_row(row: tuple[Any, ...]) -> tuple[str, ...]:
        return tuple("" if value is None else str(value) for value in row)


class CsvWorkbookReader:
    """Read CSV dimensions and retain only a bounded preview."""

    PREVIEW_ROW_LIMIT = 500

    def supports(self, file_path: Path) -> bool:
        return file_path.suffix.lower() == ".csv"

    def read_worksheets(self, file_path: Path) -> tuple[WorksheetInfo, ...]:
        row_count = 0
        column_count = 0
        headers: tuple[str, ...] = ()
        preview_rows: list[tuple[str, ...]] = []
        with file_path.open("r", encoding="utf-8-sig", newline="") as stream:
            reader = csv.reader(stream)
            for row_count, row in enumerate(reader, start=1):
                if row_count == 1:
                    headers = tuple(row)
                elif len(preview_rows) < self.PREVIEW_ROW_LIMIT:
                    preview_rows.append(tuple(row))
                column_count = max(column_count, len(row))
        return (
            WorksheetInfo(
                name=file_path.stem,
                row_count=row_count,
                column_count=column_count,
                headers=headers,
                preview_rows=tuple(preview_rows),
            ),
        )
