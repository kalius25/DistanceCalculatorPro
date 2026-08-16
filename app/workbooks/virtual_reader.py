"""Memory-bounded worksheet data sources for virtual table models."""

from __future__ import annotations

import csv
from itertools import islice
from pathlib import Path
from typing import Any, Protocol

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class VirtualWorksheetDataSource(Protocol):
    """Random-access, memory-bounded source used by a virtual table model."""

    @property
    def file_path(self) -> Path: ...

    @property
    def worksheet_name(self) -> str: ...

    @property
    def headers(self) -> tuple[str, ...]: ...

    @property
    def row_count(self) -> int: ...

    @property
    def column_count(self) -> int: ...

    def read_rows(self, start: int, count: int) -> tuple[tuple[str, ...], ...]: ...

    def close(self) -> None: ...


class UnsupportedVirtualWorkbookError(ValueError):
    """Raised when no virtual worksheet source supports a file format."""


class VirtualWorksheetNotFoundError(ValueError):
    """Raised when a requested worksheet is not present in the workbook."""


def _validate_range(start: int, count: int) -> None:
    if start < 0:
        raise ValueError("start must be non-negative")
    if count < 0:
        raise ValueError("count must be non-negative")


def _normalize_row(row: tuple[Any, ...], width: int) -> tuple[str, ...]:
    values = tuple("" if value is None else str(value) for value in row[:width])
    if len(values) < width:
        values += ("",) * (width - len(values))
    return values


class OpenPyXLVirtualWorksheetDataSource:
    """Keep one read-only Excel worksheet open and fetch bounded row ranges."""

    SUPPORTED_EXTENSIONS = frozenset({".xlsx", ".xlsm"})

    def __init__(self, file_path: str | Path, worksheet_name: str) -> None:
        self._file_path = Path(file_path)
        self._workbook: Workbook | None = load_workbook(
            filename=self._file_path,
            read_only=True,
            data_only=True,
            keep_vba=self._file_path.suffix.lower() == ".xlsm",
        )
        if worksheet_name not in self._workbook.sheetnames:
            self._workbook.close()
            self._workbook = None
            raise VirtualWorksheetNotFoundError(
                f"Worksheet not found: {worksheet_name}"
            )

        worksheet = self._workbook[worksheet_name]
        self._worksheet: Worksheet | None = worksheet
        self._worksheet_name = worksheet_name
        self._column_count = worksheet.max_column or 0
        self._row_count = max((worksheet.max_row or 0) - 1, 0)
        first_row = next(
            worksheet.iter_rows(
                min_row=1,
                max_row=1,
                max_col=self._column_count or 1,
                values_only=True,
            ),
            (),
        )
        self._headers = _normalize_row(first_row, self._column_count)

    @property
    def file_path(self) -> Path:
        return self._file_path

    @property
    def worksheet_name(self) -> str:
        return self._worksheet_name

    @property
    def headers(self) -> tuple[str, ...]:
        return self._headers

    @property
    def row_count(self) -> int:
        return self._row_count

    @property
    def column_count(self) -> int:
        return self._column_count

    def read_rows(self, start: int, count: int) -> tuple[tuple[str, ...], ...]:
        _validate_range(start, count)
        if count == 0 or start >= self._row_count:
            return ()
        worksheet = self._require_open()
        stop = min(start + count, self._row_count)
        rows = worksheet.iter_rows(
            min_row=start + 2,
            max_row=stop + 1,
            max_col=self._column_count or 1,
            values_only=True,
        )
        return tuple(_normalize_row(row, self._column_count) for row in rows)

    def close(self) -> None:
        if self._workbook is not None:
            self._workbook.close()
            self._workbook = None
            self._worksheet = None

    def _require_open(self) -> Worksheet:
        if self._worksheet is None:
            raise RuntimeError("Virtual worksheet data source is closed")
        return self._worksheet


class CsvVirtualWorksheetDataSource:
    """Scan CSV metadata once and fetch requested ranges without retaining all rows."""

    def __init__(self, file_path: str | Path) -> None:
        self._file_path = Path(file_path)
        self._worksheet_name = self._file_path.stem
        self._headers: tuple[str, ...] = ()
        self._row_count = 0
        self._column_count = 0
        self._closed = False
        self._scan_metadata()

    @property
    def file_path(self) -> Path:
        return self._file_path

    @property
    def worksheet_name(self) -> str:
        return self._worksheet_name

    @property
    def headers(self) -> tuple[str, ...]:
        return self._headers

    @property
    def row_count(self) -> int:
        return self._row_count

    @property
    def column_count(self) -> int:
        return self._column_count

    def read_rows(self, start: int, count: int) -> tuple[tuple[str, ...], ...]:
        _validate_range(start, count)
        if self._closed:
            raise RuntimeError("Virtual worksheet data source is closed")
        if count == 0 or start >= self._row_count:
            return ()

        with self._file_path.open("r", encoding="utf-8-sig", newline="") as stream:
            reader = csv.reader(stream)
            next(reader, None)
            rows = islice(reader, start, start + count)
            return tuple(_normalize_row(tuple(row), self._column_count) for row in rows)

    def close(self) -> None:
        self._closed = True

    def _scan_metadata(self) -> None:
        total_rows = 0
        with self._file_path.open("r", encoding="utf-8-sig", newline="") as stream:
            reader = csv.reader(stream)
            for total_rows, row in enumerate(reader, start=1):
                if total_rows == 1:
                    self._headers = tuple(row)
                self._column_count = max(self._column_count, len(row))
        self._row_count = max(total_rows - 1, 0)
        self._headers = _normalize_row(self._headers, self._column_count)


class VirtualWorksheetDataSourceFactory:
    """Create the virtual source appropriate for a supported workbook file."""

    def create(
        self,
        file_path: str | Path,
        worksheet_name: str | None = None,
    ) -> VirtualWorksheetDataSource:
        path = Path(file_path)
        extension = path.suffix.lower()
        if extension in OpenPyXLVirtualWorksheetDataSource.SUPPORTED_EXTENSIONS:
            if not worksheet_name:
                raise VirtualWorksheetNotFoundError(
                    "Worksheet name is required for Excel workbooks"
                )
            return OpenPyXLVirtualWorksheetDataSource(path, worksheet_name)
        if extension == ".csv":
            expected_name = path.stem
            if worksheet_name is not None and worksheet_name != expected_name:
                raise VirtualWorksheetNotFoundError(
                    f"Worksheet not found: {worksheet_name}"
                )
            return CsvVirtualWorksheetDataSource(path)
        raise UnsupportedVirtualWorkbookError(
            f"Unsupported virtual workbook format: {extension}"
        )


__all__ = [
    "CsvVirtualWorksheetDataSource",
    "OpenPyXLVirtualWorksheetDataSource",
    "UnsupportedVirtualWorkbookError",
    "VirtualWorksheetDataSource",
    "VirtualWorksheetDataSourceFactory",
    "VirtualWorksheetNotFoundError",
]
